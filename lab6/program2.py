import csv
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def calculate_age(birth_date_str):
    try:
        birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y')
        today = datetime.now()
        age = today.year - birth_date.year
        
        if (today.month, today.day) < (birth_date.month, birth_date.day):
            age -= 1
        
        return age
    except Exception as e:
        print(f"Помилка при обчисленні віку для дати {birth_date_str}: {e}")
        return 0


def get_age_category(age):
    if age < 18:
        return "younger_18"
    elif age <= 45:
        return "18-45"
    elif age <= 70:
        return "45-70"
    else:
        return "older_70"


def style_worksheet(ws, is_all_sheet=False):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def read_csv_and_create_xlsx(csv_filename='employees.csv', xlsx_filename='employees.xlsx'):
    csv_path = Path(csv_filename)
    if not csv_path.exists():
        print(f" Помилка: Файл '{csv_filename}' не знайдено!")
        print(f"   Перевірте, чи файл знаходиться у поточній директорії.")
        return False
    
    try:
        print(f" Читання CSV файлу '{csv_filename}'...")
        employees = []
        
        with open(csv_filename, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                age = calculate_age(row['Дата народження'])
                row['Вік'] = age
                employees.append(row)
        
        print(f"   Прочитано {len(employees)} записів")
        
    except Exception as e:
        print(f" Помилка при відкритті або читанні CSV файлу: {e}")
        return False
    
    try:
        print(f"\n Створення XLSX файлу '{xlsx_filename}'...")
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        sheet_names = ["all", "younger_18", "18-45", "45-70", "older_70"]
        sheets = {}
        
        for name in sheet_names:
            sheets[name] = wb.create_sheet(title=name)
        
        all_headers = ['Прізвище', 'Імя', 'По бат', 'Стать', 'Дата народження', 
                       'Посада', 'Місто проживання', 'Адреса прож', 'Телефон', 'Email']
        
        age_headers = ['№', 'Прізвище', 'Імя', 'По бат', 'Дата народження', 'Вік']
        
        sheets["all"].append(all_headers)
        for emp in employees:
            row_data = [emp.get(header, '') for header in all_headers]
            sheets["all"].append(row_data)
        
        style_worksheet(sheets["all"], is_all_sheet=True)
        
        category_counters = {
            "younger_18": 1,
            "18-45": 1,
            "45-70": 1,
            "older_70": 1
        }
        
        for category in ["younger_18", "18-45", "45-70", "older_70"]:
            sheets[category].append(age_headers)
        
        for emp in employees:
            age = emp['Вік']
            category = get_age_category(age)
            
            row_data = [
                category_counters[category],
                emp['Прізвище'],
                emp['Імя'],
                emp['По бат'],
                emp['Дата народження'],
                age
            ]
            
            sheets[category].append(row_data)
            category_counters[category] += 1
        
        for category in ["younger_18", "18-45", "45-70", "older_70"]:
            style_worksheet(sheets[category], is_all_sheet=False)
        
        wb.save(xlsx_filename)
        
        print(f"\n Статистика розподілу за віковими категоріями:")
        print(f"   • До 18 років (younger_18): {category_counters['younger_18'] - 1} осіб")
        print(f"   • 18-45 років (18-45): {category_counters['18-45'] - 1} осіб")
        print(f"   • 45-70 років (45-70): {category_counters['45-70'] - 1} осіб")
        print(f"   • Старше 70 років (older_70): {category_counters['older_70'] - 1} осіб")
        print(f"\n Ok")
        print(f"   Файл '{xlsx_filename}' успішно створено!")
        
        return True
        
    except Exception as e:
        print(f" Помилка при створенні XLSX файлу: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("=" * 70)
    print("Програма 2 - Створення XLSX файлу з віковими категоріями")
    print("=" * 70)
    print()
    
    success = read_csv_and_create_xlsx('employees.csv', 'employees.xlsx')
    
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
